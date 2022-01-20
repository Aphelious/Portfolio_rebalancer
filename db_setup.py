# -*- coding: utf-8 -*-

import sqlite3
from env import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime



#Connect to database

conn = sqlite3.connect('db_file')
cursor = conn.cursor()


#Create tables

# cursor.execute('''
# DROP TABLE transactions
# ''')
# conn.commit()

cursor.execute('''CREATE TABLE transactions (
transactions_id INTEGER PRIMARY KEY,
account INTEGER,
trade_date TEXT,
settlement_date TEXT,
trans_type TEXT,
investment_name TEXT,
category TEXT,
ticker TEXT,
shares REAL,
share_price REAL,
amount REAL,
principal_amount REAL);
''')
conn.commit()


# Connect and read from the spreadsheet

wb = load_workbook(f'{path}{csv_filename}')
ws = wb.active
wb.iso_dates = True

for row in range(2, 88):
    data = []
    data.append(row-1)
    for col in range(1,12):
        char = get_column_letter(col)
        cell_value = ws[f'{char}{row}'].value
        if type(cell_value) == str and '\xa0' in cell_value:
            cell_value = cell_value.replace('\xa0', ' ')
        if type(cell_value) == datetime.datetime:
            cell_value = str(cell_value)
        data.append(cell_value)
    data = tuple(data)
    print(data)
    cursor.execute(f'''INSERT INTO transactions VALUES{data}''')

conn.commit()


# #Simple SQLite shell:

# buffer = ""
#
# print("Enter your SQL commands to execute in sqlite3.")
# print("Enter a blank line to exit.")
#
# while True:
#     line = input()
#     if line == "":
#         break
#     buffer += line
#     if sqlite3.complete_statement(buffer):
#         try:
#             buffer = buffer.strip()
#             cursor.execute(buffer)
#             if buffer.lstrip().upper().startswith("SELECT"):
#                 result = cursor.fetchall()
#                 if result is []:
#                     print('Result set is empty')
#                 else:
#                     for i in result:
#                         print(i)
#         except sqlite3.Error as e:
#             print("An error occurred:", e.args[0])
#         buffer = ""


#Print out the table

# cursor.execute('''
# SELECT * FROM transactions
# WHERE Category = 'Balanced';
# ''')
# result = cursor.fetchall()
# col_names = []
# for i in cursor.description:
#     col_names.append(i[0])
# print(col_names)
# if result is []:
#     print('Result set is empty')
# else:
#     for i in result:
#         print(i)




#Check rows affected and close the db connection

def check_rows():
    if cursor.rowcount == -1:
        return 0
    else:
        return cursor.rowcount

print(f'Number of rows modified: {check_rows()}')
conn.close()
